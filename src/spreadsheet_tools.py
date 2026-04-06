import json
import logging
import os
import re
import subprocess
import sys
from copy import copy
from datetime import datetime
from typing import Any

import file_tools

logger = logging.getLogger("talos.spreadsheet_tools")

try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except Exception:
    pd = None
    PANDAS_AVAILABLE = False

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Color

    OPENPYXL_AVAILABLE = True
except Exception:
    openpyxl = None
    Workbook = None
    Color = None
    OPENPYXL_AVAILABLE = False

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_EXTERNAL_LINK_RE = re.compile(r"\[[^\]]+\]")

FORMULA_ERROR_TOKENS = {
    "#REF!",
    "#DIV/0!",
    "#VALUE!",
    "#NAME?",
    "#NUM!",
    "#N/A",
    "#NULL!",
}

# Financial modeling defaults
INPUT_BLUE_RGB = "FF0000FF"
FORMULA_BLACK_RGB = "FF000000"
EXTERNAL_LINK_RED_RGB = "FFFF0000"


def _error(message: str, hint: str = "") -> dict:
    payload = {"ok": False, "error": message}
    if hint:
        payload["hint"] = hint
    return payload


def _resolve_path(path: str, must_exist: bool = True) -> tuple[str | None, dict | None]:
    resolved, err = file_tools._validate_path(path)
    if err:
        return None, _error(err)

    if must_exist and not os.path.isfile(resolved):
        return None, _error(f"Workbook not found: {path}")

    return resolved, None


def _json_safe(value: Any) -> Any:
    if value is None:
        return None

    if PANDAS_AVAILABLE:
        try:
            if pd.isna(value):
                return None
        except Exception:
            pass

    if isinstance(value, (str, int, float, bool)):
        return value

    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            return str(value)

    return str(value)


def _ensure_sheet(workbook: Any, sheet_name: str | None) -> Any:
    if sheet_name and sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    if sheet_name and sheet_name not in workbook.sheetnames:
        return workbook.create_sheet(sheet_name)
    return workbook.active


def _set_font_color(cell: Any, rgb: str) -> None:
    font = copy(cell.font)
    font.color = Color(rgb=rgb)
    cell.font = font


def read_with_pandas(path: str, sheet_name: str | None = None, max_rows: int = 200, max_cols: int = 40) -> dict:
    if not PANDAS_AVAILABLE:
        return _error(
            "pandas is not installed",
            "Install dependencies: pip install -r requirements.txt",
        )

    resolved, err = _resolve_path(path, must_exist=True)
    if err:
        return err

    try:
        requested = sheet_name if sheet_name and str(sheet_name).strip().lower() != "none" else None
        data = pd.read_excel(
            resolved,
            sheet_name=requested,
            header=None,
            dtype=object,
            na_filter=False,
            engine="openpyxl",
        )

        if not isinstance(data, dict):
            data = {sheet_name or "Sheet1": data}

        sheets_payload = {}
        for name, frame in data.items():
            rows_total = int(frame.shape[0])
            cols_total = int(frame.shape[1])
            trimmed = frame.iloc[: max(1, int(max_rows)), : max(1, int(max_cols))]
            preview = []
            for _, row in trimmed.iterrows():
                preview.append([_json_safe(v) for v in row.tolist()])

            sheets_payload[str(name)] = {
                "rows_total": rows_total,
                "cols_total": cols_total,
                "preview_rows": len(preview),
                "preview_cols": min(cols_total, max(1, int(max_cols))),
                "preview": preview,
            }

        return {
            "ok": True,
            "path": resolved,
            "loaded_at": datetime.utcnow().isoformat() + "Z",
            "sheets": sheets_payload,
            "note": "Workbook loaded with pandas (full workbook loaded in memory; response is preview-limited).",
        }

    except Exception as e:
        logger.exception("read_with_pandas failed: %s", path)
        return _error(str(e))


def edit_with_openpyxl(path: str, operations: list[dict], create_if_missing: bool = False) -> dict:
    if not OPENPYXL_AVAILABLE:
        return _error(
            "openpyxl is not installed",
            "Install dependencies: pip install -r requirements.txt",
        )

    if not isinstance(operations, list) or not operations:
        return _error("operations must be a non-empty list")

    resolved, err = _resolve_path(path, must_exist=False)
    if err:
        return err

    try:
        workbook_exists = os.path.isfile(resolved)
        if workbook_exists:
            wb = openpyxl.load_workbook(resolved)
        else:
            if not create_if_missing:
                return _error("Workbook not found", "Set create_if_missing=true to create a new workbook")
            wb = Workbook()

        applied = []
        for index, op in enumerate(operations, start=1):
            if not isinstance(op, dict):
                return _error(f"operations[{index}] must be an object")

            action = str(op.get("action", "set_cell")).strip().lower()
            if action == "ensure_sheet":
                title = str(op.get("sheet", "")).strip()
                if not title:
                    return _error(f"operations[{index}] ensure_sheet requires 'sheet'")
                _ensure_sheet(wb, title)
                applied.append({"index": index, "action": action, "sheet": title})
                continue

            sheet = _ensure_sheet(wb, str(op.get("sheet", "")).strip() or None)

            if action in {"set_cell", "set_formula"}:
                address = str(op.get("cell", "")).strip()
                if not address:
                    return _error(f"operations[{index}] {action} requires 'cell'")

                value = op.get("value")
                if action == "set_formula":
                    formula = str(value if value is not None else "").strip()
                    if formula and not formula.startswith("="):
                        formula = "=" + formula
                    value = formula
                elif bool(op.get("formula", False)) and value is not None:
                    text = str(value)
                    if text and not text.startswith("="):
                        value = "=" + text

                sheet[address] = value

                number_format = op.get("number_format")
                if number_format is not None:
                    sheet[address].number_format = str(number_format)

                applied.append({"index": index, "action": action, "sheet": sheet.title, "cell": address})
                continue

            if action == "append_row":
                values = op.get("values", [])
                if not isinstance(values, list):
                    return _error(f"operations[{index}] append_row requires 'values' as array")
                sheet.append(values)
                applied.append({"index": index, "action": action, "sheet": sheet.title, "values": len(values)})
                continue

            return _error(f"Unsupported operation action: {action}")

        os.makedirs(os.path.dirname(resolved), exist_ok=True)
        wb.save(resolved)

        return {
            "ok": True,
            "path": resolved,
            "created": not workbook_exists,
            "operations_applied": len(applied),
            "operations": applied,
            "note": "Workbook updated with openpyxl to preserve formulas and rich formatting.",
        }

    except Exception as e:
        logger.exception("edit_with_openpyxl failed: %s", path)
        return _error(str(e))


def verify_formula_errors(path: str, sheet_name: str | None = None, max_errors: int = 500) -> dict:
    if not OPENPYXL_AVAILABLE:
        return _error(
            "openpyxl is not installed",
            "Install dependencies: pip install -r requirements.txt",
        )

    resolved, err = _resolve_path(path, must_exist=True)
    if err:
        return err

    try:
        wb = openpyxl.load_workbook(resolved, data_only=False, read_only=True)
        target_sheets = [sheet_name] if sheet_name else list(wb.sheetnames)

        findings = []
        per_sheet_counts = {}

        for sheet in target_sheets:
            if sheet not in wb.sheetnames:
                return _error(f"Sheet not found: {sheet}")
            ws = wb[sheet]
            count = 0

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    value = cell.value
                    if value is None:
                        continue

                    is_error = False
                    detail = ""

                    if getattr(cell, "data_type", "") == "e":
                        is_error = True
                        detail = "error_cell"
                    elif isinstance(value, str):
                        upper = value.strip().upper()
                        if upper in FORMULA_ERROR_TOKENS:
                            is_error = True
                            detail = "error_literal"
                        elif upper.startswith("="):
                            for token in FORMULA_ERROR_TOKENS:
                                if token in upper:
                                    is_error = True
                                    detail = "formula_contains_error_token"
                                    break

                    if is_error:
                        count += 1
                        findings.append(
                            {
                                "sheet": ws.title,
                                "cell": cell.coordinate,
                                "value": str(value),
                                "reason": detail,
                            }
                        )
                        if len(findings) >= max(1, int(max_errors)):
                            break
                if len(findings) >= max(1, int(max_errors)):
                    break

            per_sheet_counts[ws.title] = count
            if len(findings) >= max(1, int(max_errors)):
                break

        return {
            "ok": True,
            "path": resolved,
            "error_count": len(findings),
            "errors": findings,
            "per_sheet_counts": per_sheet_counts,
            "has_errors": len(findings) > 0,
            "summary": "zero formula errors detected" if len(findings) == 0 else "formula errors detected",
        }

    except Exception as e:
        logger.exception("verify_formula_errors failed: %s", path)
        return _error(str(e))


def _collect_input_coords(ws: Any, ranges: list[str]) -> set[str]:
    coords: set[str] = set()
    for raw in ranges:
        rng = str(raw or "").strip()
        if not rng:
            continue
        try:
            for row in ws[rng]:
                for cell in row:
                    coords.add(cell.coordinate)
        except Exception:
            continue
    return coords


def apply_financial_color_coding(
    path: str,
    sheet_name: str | None = None,
    input_ranges: list[str] | None = None,
    header_rows: int = 1,
) -> dict:
    if not OPENPYXL_AVAILABLE:
        return _error(
            "openpyxl is not installed",
            "Install dependencies: pip install -r requirements.txt",
        )

    resolved, err = _resolve_path(path, must_exist=True)
    if err:
        return err

    try:
        wb = openpyxl.load_workbook(resolved)
        target_sheets = [sheet_name] if sheet_name else list(wb.sheetnames)
        if sheet_name and sheet_name not in wb.sheetnames:
            return _error(f"Sheet not found: {sheet_name}")

        changes = {
            "inputs_blue": 0,
            "formulas_black": 0,
            "external_links_red": 0,
        }

        header_rows = max(0, int(header_rows))
        selected_ranges = input_ranges if isinstance(input_ranges, list) else []

        for name in target_sheets:
            ws = wb[name]
            explicit_input_coords = _collect_input_coords(ws, selected_ranges)

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    value = cell.value
                    if value is None:
                        continue

                    if isinstance(value, str) and value.startswith("="):
                        if _EXTERNAL_LINK_RE.search(value) or "http://" in value.lower() or "https://" in value.lower():
                            _set_font_color(cell, EXTERNAL_LINK_RED_RGB)
                            changes["external_links_red"] += 1
                        else:
                            _set_font_color(cell, FORMULA_BLACK_RGB)
                            changes["formulas_black"] += 1
                        continue

                    should_color_input = False
                    if explicit_input_coords:
                        should_color_input = cell.coordinate in explicit_input_coords
                    else:
                        should_color_input = cell.row > header_rows

                    if should_color_input:
                        _set_font_color(cell, INPUT_BLUE_RGB)
                        changes["inputs_blue"] += 1

        wb.save(resolved)

        return {
            "ok": True,
            "path": resolved,
            "applied": changes,
            "legend": {
                "inputs": "blue",
                "formulas": "black",
                "external_links": "red",
            },
        }

    except Exception as e:
        logger.exception("apply_financial_color_coding failed: %s", path)
        return _error(str(e))


def recalculate_with_libreoffice(
    path: str,
    output_path: str | None = None,
    timeout_s: int = 180,
    script_path: str | None = None,
) -> dict:
    resolved, err = _resolve_path(path, must_exist=True)
    if err:
        return err

    if output_path:
        resolved_output, output_err = _resolve_path(output_path, must_exist=False)
        if output_err:
            return output_err
    else:
        resolved_output = ""

    if script_path:
        recalc_script = script_path if os.path.isabs(script_path) else os.path.join(_SCRIPT_DIR, script_path)
    else:
        recalc_script = os.path.join(_SCRIPT_DIR, "scripts", "recalc.py")

    if not os.path.isfile(recalc_script):
        return _error(f"Recalc script not found: {recalc_script}")

    cmd = [sys.executable, recalc_script, "--input", resolved, "--timeout", str(max(30, int(timeout_s)))]
    if resolved_output:
        cmd.extend(["--output", resolved_output])

    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=max(45, int(timeout_s) + 15),
        )

        stdout = (result.stdout or "").strip()
        stderr = (result.stderr or "").strip()

        if result.returncode != 0:
            return _error(
                "LibreOffice recalculation failed",
                hint=(stderr or stdout or "Check LibreOffice installation and script logs."),
            )

        parsed = None
        if stdout:
            lines = [line.strip() for line in stdout.splitlines() if line.strip()]
            if lines:
                try:
                    parsed = json.loads(lines[-1])
                except Exception:
                    parsed = None

        if isinstance(parsed, dict):
            return parsed

        return {
            "ok": True,
            "path": resolved,
            "output_path": resolved_output or resolved,
            "note": "Recalculation command completed. Verify formula errors with verify_formula_errors.",
        }

    except subprocess.TimeoutExpired:
        return _error("LibreOffice recalculation timed out")
    except Exception as e:
        logger.exception("recalculate_with_libreoffice failed: %s", path)
        return _error(str(e))


def execute(action: str, **kwargs) -> dict:
    normalized = str(action or "").strip().lower()

    if normalized in {"read_with_pandas", "read", "preview"}:
        return read_with_pandas(
            path=str(kwargs.get("path", "")).strip(),
            sheet_name=(str(kwargs.get("sheet_name", "")).strip() or None),
            max_rows=kwargs.get("max_rows", 200),
            max_cols=kwargs.get("max_cols", 40),
        )

    if normalized in {"edit_with_openpyxl", "edit", "update"}:
        return edit_with_openpyxl(
            path=str(kwargs.get("path", "")).strip(),
            operations=kwargs.get("operations", []),
            create_if_missing=bool(kwargs.get("create_if_missing", False)),
        )

    if normalized in {"recalculate_with_libreoffice", "recalc", "recalculate"}:
        return recalculate_with_libreoffice(
            path=str(kwargs.get("path", "")).strip(),
            output_path=(str(kwargs.get("output_path", "")).strip() or None),
            timeout_s=kwargs.get("timeout_s", 180),
            script_path=(str(kwargs.get("script_path", "")).strip() or None),
        )

    if normalized in {"verify_formula_errors", "verify_errors", "validate"}:
        return verify_formula_errors(
            path=str(kwargs.get("path", "")).strip(),
            sheet_name=(str(kwargs.get("sheet_name", "")).strip() or None),
            max_errors=kwargs.get("max_errors", 500),
        )

    if normalized in {"apply_financial_color_coding", "apply_colors", "financial_colors"}:
        return apply_financial_color_coding(
            path=str(kwargs.get("path", "")).strip(),
            sheet_name=(str(kwargs.get("sheet_name", "")).strip() or None),
            input_ranges=kwargs.get("input_ranges"),
            header_rows=kwargs.get("header_rows", 1),
        )

    return _error(
        f"Unsupported spreadsheet action: {action}",
        hint=(
            "Supported actions: read_with_pandas, edit_with_openpyxl, "
            "recalculate_with_libreoffice, verify_formula_errors, apply_financial_color_coding"
        ),
    )
